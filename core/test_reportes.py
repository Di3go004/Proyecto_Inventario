"""
Pruebas de los reportes (RF-14).

Lo que más importa: que lo que se descarga coincida con lo que se está
viendo en pantalla. Es el error clásico de los reportes — cada camino hace
su propia consulta y con el tiempo dejan de cuadrar. Acá los dos usan
core/reportes.py, y estas pruebas lo comprueban.
"""

import io
from datetime import timedelta

import openpyxl
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone

from core import exportar, reportes
from core.models import Bodega
from tecnica.ayuda_pruebas import dar_de_baja, dar_existencia
from tecnica.models import Activo, PrestamoActivo
from usuarios.models import Usuario
from ventas.models import Articulo, MovimientoVenta


class BaseReportes(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.b1 = Bodega.objects.create(nombre='Bodega 1', tipo=Bodega.Tipo.VENTA)
        cls.b2 = Bodega.objects.create(nombre='Bodega 2', tipo=Bodega.Tipo.VENTA)
        cls.btec = Bodega.objects.create(nombre='Bodega Técnica', tipo=Bodega.Tipo.TECNICA)

        cls.admin = Usuario.objects.create_user(
            username='admin_rep', password='clave-de-prueba', rol=Usuario.Rol.ADMINISTRADOR,
        )
        cls.operador = Usuario.objects.create_user(
            username='op_rep', password='clave-de-prueba', rol=Usuario.Rol.OPERADOR,
        )
        cls.contable = Usuario.objects.create_user(
            username='cont_rep', password='clave-de-prueba', rol=Usuario.Rol.CONTABILIDAD,
        )

        # Bodega 1: uno óptimo (30) y uno crítico (0)
        cls.optimo = cls.articulo('Báscula grande', 'BG-1', cls.b1, precio=100)
        cls.critico = cls.articulo('Adaptador', 'AD-1', cls.b1, precio=50)
        # Bodega 2: uno en alerta (4)
        cls.alerta = cls.articulo('Celda de carga', 'CC-1', cls.b2, precio=200)

        cls.ingresar(cls.optimo, 30)
        cls.ingresar(cls.alerta, 4)

        cls.taladro = Activo.objects.create(
            codigo_interno='SE-T1', nombre_producto='Taladro', bodega=cls.btec, precio=900,
        )
        cls.agotado = Activo.objects.create(
            codigo_interno='SE-T2', nombre_producto='Sierra vieja', bodega=cls.btec, precio=500,
        )
        # Un taladro entra y se queda; de la sierra entró una y se dio de baja.
        dar_existencia(cls.taladro, 1, cls.admin)
        dar_existencia(cls.agotado, 1, cls.admin)
        dar_de_baja(cls.agotado, 1, cls.admin)

    @classmethod
    def articulo(cls, nombre, modelo, bodega, precio):
        return Articulo.objects.create(
            nombre_producto=nombre, modelo=modelo, capacidad='1kg',
            bodega=bodega, precio=precio,
        )

    @classmethod
    def ingresar(cls, articulo, cantidad, cuando=None):
        return MovimientoVenta.objects.create(
            folio='ING-TEST', articulo=articulo, cantidad=cantidad,
            tipo_documento=MovimientoVenta.TipoDocumento.INGRESO,
            tipo_transaccion=MovimientoVenta.TipoTransaccion.VENTA,
            fecha=cuando or timezone.now(), usuario=cls.admin,
        )

    def setUp(self):
        self.client.force_login(self.admin)

    def hoja(self, url, **parametros):
        parametros['formato'] = 'excel'
        respuesta = self.client.get(url, parametros)
        self.assertEqual(respuesta.status_code, 200)
        return openpyxl.load_workbook(io.BytesIO(respuesta.content)).active

    def fila_del_encabezado(self, hoja, primer_titulo):
        """
        Busca en qué fila cayó de verdad el texto del encabezado, en vez de
        confiar en la constante. Si las pruebas usaran la constante, moverla
        movería también lo que comprueban y no detectarían un desfase.
        """
        for numero in range(1, 15):
            if hoja.cell(row=numero, column=1).value == primer_titulo:
                return numero
        self.fail(f'no se encontró el encabezado "{primer_titulo}" en la hoja')

    def columna(self, hoja, fila_encabezado, titulo):
        """
        Busca la columna por su título y no por su número. Insertar una columna
        nueva corre todas las que van después: con el número escrito a mano,
        la prueba pasaría a mirar la columna de al lado sin avisar de nada.
        """
        for numero in range(1, 25):
            if hoja.cell(row=fila_encabezado, column=numero).value == titulo:
                return numero
        self.fail(f'no se encontró la columna "{titulo}" en la hoja')


class AccesoTests(BaseReportes):
    RUTAS = ['indice_reportes', 'reporte_existencias', 'reporte_alertas',
             'reporte_movimientos', 'reporte_prestamos']

    def test_los_ven_administrador_y_contabilidad(self):
        """RF-04: para contabilidad los reportes son la razón de su acceso."""
        for usuario in (self.admin, self.contable):
            self.client.force_login(usuario)
            for nombre in self.RUTAS:
                with self.subTest(rol=usuario.rol, pantalla=nombre):
                    self.assertEqual(self.client.get(reverse(nombre)).status_code, 200)

    def test_el_operador_no(self):
        """
        Antes los veían los tres roles. El operador mueve bodega —entradas,
        salidas, préstamos— y para eso no necesita saber cuánto vale el
        inventario, que es información de la empresa que no le toca.

        Lo que sí necesita —qué reponer y qué está prestado— lo sigue viendo
        en el Resumen. Ver usuarios/test_operador.py.
        """
        self.client.force_login(self.operador)
        for nombre in self.RUTAS:
            with self.subTest(pantalla=nombre):
                self.assertEqual(self.client.get(reverse(nombre)).status_code, 403)

    def test_contabilidad_tambien_puede_descargar(self):
        self.client.force_login(self.contable)
        respuesta = self.client.get(reverse('reporte_existencias'), {'formato': 'excel'})
        self.assertEqual(respuesta.status_code, 200)

    def test_sin_sesion_no_se_entra(self):
        self.client.logout()
        for nombre in self.RUTAS:
            with self.subTest(pantalla=nombre):
                respuesta = self.client.get(reverse(nombre))
                self.assertEqual(respuesta.status_code, 302)
                self.assertIn('/login/', respuesta.url)


class ExistenciasTests(BaseReportes):
    def test_agrupa_y_valoriza_por_bodega(self):
        filas, _detalle, totales = reportes.existencias()
        por_nombre = {f['bodega']: f for f in filas}

        # Bodega 1: 30 unidades del de Q100 + 0 del de Q50 = Q3,000
        self.assertEqual(por_nombre['Bodega 1']['unidades'], 30)
        self.assertEqual(por_nombre['Bodega 1']['valor'], 3000)
        # Bodega 2: 4 unidades de Q200 = Q800
        self.assertEqual(por_nombre['Bodega 2']['valor'], 800)
        self.assertEqual(totales['valor'], 3800)

    def test_cuenta_los_niveles_de_alerta(self):
        filas, _d, _t = reportes.existencias()
        b1 = next(f for f in filas if f['bodega'] == 'Bodega 1')

        self.assertEqual(b1['niveles']['optimo'], 1)
        self.assertEqual(b1['niveles']['critico'], 1)

    def test_la_bodega_tecnica_no_cuenta_lo_que_se_dio_de_baja(self):
        """Lo descartado queda en existencia 0 y deja de valer (RF-12)."""
        tecnica = reportes.valorizacion_tecnica()

        self.assertEqual(tecnica['valor'], 900, 'la sierra dada de baja no suma')
        self.assertEqual(tecnica['unidades'], 1)
        self.assertEqual(tecnica['agotados'], 1)

    def test_la_valorizacion_tecnica_multiplica_por_la_existencia(self):
        """
        Regresión: sumaba solo los precios porque se asumía una unidad por
        registro. Con 10 bombillos de Q25 la bodega salía valorizada en 25.
        """
        bombillos = Activo.objects.create(
            codigo_interno='SE-T3', nombre_producto='Bombillo', bodega=self.btec, precio=25,
        )
        dar_existencia(bombillos, 10, self.admin)

        tecnica = reportes.valorizacion_tecnica()

        self.assertEqual(tecnica['valor'], 900 + 250)
        self.assertEqual(tecnica['unidades'], 11)

    def test_se_puede_filtrar_por_bodega(self):
        filas, _d, totales = reportes.existencias(bodega_id=self.b2.pk)

        self.assertEqual(len(filas), 1)
        self.assertEqual(totales['valor'], 800)

    def test_los_inactivos_quedan_fuera_por_defecto(self):
        Articulo.objects.filter(pk=self.alerta.pk).update(activo=False)

        _f, _d, con = reportes.existencias(solo_activos=True)
        _f2, _d2, sin = reportes.existencias(solo_activos=False)

        self.assertEqual(con['valor'], 3000)
        self.assertEqual(sin['valor'], 3800)


class AlertasTests(BaseReportes):
    def test_solo_trae_lo_que_esta_bajo_el_umbral(self):
        encontrados = reportes.alertas_de_stock()
        codigos = {a.pk for a in encontrados}

        self.assertIn(self.critico.pk, codigos)
        self.assertIn(self.alerta.pk, codigos)
        self.assertNotIn(self.optimo.pk, codigos, 'el que está óptimo no es una alerta')

    def test_lo_mas_urgente_va_primero(self):
        encontrados = reportes.alertas_de_stock()
        self.assertEqual(encontrados[0].pk, self.critico.pk, 'el que está en cero primero')

    def test_no_incluye_articulos_inactivos(self):
        Articulo.objects.filter(pk=self.critico.pk).update(activo=False)
        self.assertNotIn(self.critico.pk, {a.pk for a in reportes.alertas_de_stock()})


class MovimientosTests(BaseReportes):
    def test_separa_lo_que_entro_de_lo_que_salio(self):
        MovimientoVenta.objects.create(
            folio='SAL-TEST', articulo=self.optimo, cantidad=5,
            tipo_documento=MovimientoVenta.TipoDocumento.SALIDA,
            tipo_transaccion=MovimientoVenta.TipoTransaccion.VENTA,
            fecha=timezone.now(), usuario=self.admin,
        )

        resumen, _detalle = reportes.movimientos_del_periodo()

        self.assertEqual(resumen['ingresos'], 34)
        self.assertEqual(resumen['salidas'], 5)

    def test_respeta_el_rango_de_fechas(self):
        viejo = timezone.now() - timedelta(days=40)
        self.ingresar(self.critico, 7, cuando=viejo)

        resumen, _d = reportes.movimientos_del_periodo(
            desde=(timezone.now() - timedelta(days=7)).date(),
        )

        self.assertEqual(resumen['ingresos'], 34, 'el de hace 40 días queda fuera')

    def test_cuenta_los_documentos_distintos(self):
        resumen, _d = reportes.movimientos_del_periodo()
        self.assertEqual(resumen['documentos'], 1, 'los dos ingresos comparten folio')


class PrestamosTests(BaseReportes):
    def test_junta_los_dos_modulos(self):
        MovimientoVenta.objects.create(
            folio='SAL-DEMO', articulo=self.optimo, cantidad=1,
            tipo_documento=MovimientoVenta.TipoDocumento.SALIDA,
            tipo_transaccion=MovimientoVenta.TipoTransaccion.PRESTAMO_DEMO,
            cliente_nombre='Cliente Demo', fecha=timezone.now(), usuario=self.admin,
        )
        PrestamoActivo.objects.create(
            activo=self.taladro, solicitante='Carlos',
            estado_al_salir=Activo.Estado.BUEN_ESTADO, usuario=self.admin,
        )

        filas = reportes.prestamos_abiertos()
        origenes = {f['origen'] for f in filas}

        self.assertEqual(len(filas), 2)
        self.assertEqual(origenes, {'Bodega 1 y 2', 'Bodega Técnica'})

    def test_lo_que_lleva_mas_tiempo_afuera_va_primero(self):
        antiguo = PrestamoActivo.objects.create(
            activo=self.taladro, solicitante='Hace mucho',
            estado_al_salir=Activo.Estado.BUEN_ESTADO, usuario=self.admin,
        )
        PrestamoActivo.objects.filter(pk=antiguo.pk).update(
            fecha_salida=timezone.now() - timedelta(days=45),
        )
        MovimientoVenta.objects.create(
            folio='SAL-HOY', articulo=self.optimo, cantidad=1,
            tipo_documento=MovimientoVenta.TipoDocumento.SALIDA,
            tipo_transaccion=MovimientoVenta.TipoTransaccion.PRESTAMO_DEMO,
            fecha=timezone.now(), usuario=self.admin,
        )

        filas = reportes.prestamos_abiertos()

        self.assertEqual(filas[0]['quien'], 'Hace mucho')
        self.assertGreaterEqual(filas[0]['dias'], 45)

    def test_lo_devuelto_ya_no_aparece(self):
        prestamo = PrestamoActivo.objects.create(
            activo=self.taladro, solicitante='Carlos',
            estado_al_salir=Activo.Estado.BUEN_ESTADO, usuario=self.admin,
        )
        prestamo.fecha_regreso = timezone.now()
        prestamo.recibido_por = 'Bodega'
        prestamo.estado_al_regresar = Activo.Estado.BUEN_ESTADO
        prestamo.save()

        self.assertEqual(reportes.prestamos_abiertos(), [])


class ExcelTests(BaseReportes):
    def test_los_cuatro_generan_un_archivo_valido(self):
        rutas = ['reporte_existencias', 'reporte_alertas',
                 'reporte_movimientos', 'reporte_prestamos']
        for nombre in rutas:
            with self.subTest(reporte=nombre):
                respuesta = self.client.get(reverse(nombre), {'formato': 'excel'})
                self.assertEqual(respuesta.status_code, 200)
                self.assertIn('spreadsheetml', respuesta['Content-Type'])
                self.assertIn('attachment', respuesta['Content-Disposition'])
                self.assertEqual(respuesta.content[:2], b'PK', 'un .xlsx es un zip')

    def test_el_formato_cae_en_la_misma_fila_que_el_texto_del_encabezado(self):
        """
        Regresión: la cabecera se escribía con append() después de una fila
        vacía, y como appendear una lista vacía adelanta el cursor de openpyxl
        sin crear celdas, max_row quedaba atrás. El relleno azul terminaba
        pintado en la fila en blanco, el panel fijo una fila más arriba y el
        filtro dejando fuera la última fila de datos.

        No se compara contra la constante a propósito: se busca dónde cayó el
        texto de verdad, que es lo que el error hacía divergir.
        """
        hoja = self.hoja(reverse('reporte_existencias'))
        fila = self.fila_del_encabezado(hoja, 'Código')
        encabezado = hoja.cell(row=fila, column=1)

        self.assertTrue(encabezado.font.bold, 'el encabezado va en negrita')
        self.assertIn(exportar.AZUL, str(encabezado.fill.fgColor.rgb))
        self.assertEqual(hoja.freeze_panes, f'A{fila + 1}')
        self.assertTrue(hoja.auto_filter.ref.startswith(f'A{fila}:'))

    def test_el_filtro_cubre_hasta_la_ultima_fila(self):
        hoja = self.hoja(reverse('reporte_existencias'))
        self.assertTrue(
            hoja.auto_filter.ref.endswith(str(hoja.max_row)),
            f'el filtro {hoja.auto_filter.ref} no llega a la fila {hoja.max_row}',
        )

    def test_los_montos_van_como_numero_y_no_como_texto(self):
        """Si llegan como texto, en Excel no se pueden sumar ni ordenar."""
        hoja = self.hoja(reverse('reporte_existencias'))
        fila = self.fila_del_encabezado(hoja, 'Código')
        precio = hoja.cell(row=fila + 1, column=self.columna(hoja, fila, 'Precio unitario'))

        self.assertIsInstance(precio.value, (int, float))
        self.assertEqual(precio.number_format, exportar.FORMATO_MONEDA)

    def test_las_fechas_van_como_fecha(self):
        hoja = self.hoja(reverse('reporte_movimientos'))
        fila = self.fila_del_encabezado(hoja, 'Fecha')

        self.assertEqual(hoja.cell(row=fila + 1, column=1).number_format, exportar.FORMATO_FECHA)

    def test_lo_descargado_coincide_con_lo_que_se_ve(self):
        """
        El punto de que pantalla y exportación compartan core/reportes.py: si
        cada una consultara por su cuenta, tarde o temprano dejarían de cuadrar.
        """
        pantalla = self.client.get(reverse('reporte_existencias'))
        en_pantalla = pantalla.context['totales']['articulos']

        hoja = self.hoja(reverse('reporte_existencias'))
        en_excel = hoja.max_row - self.fila_del_encabezado(hoja, 'Código')

        self.assertEqual(en_excel, en_pantalla)

    def test_el_filtro_de_la_pantalla_se_aplica_a_la_descarga(self):
        hoja = self.hoja(reverse('reporte_existencias'), bodega=self.b2.pk)
        filas = hoja.max_row - self.fila_del_encabezado(hoja, 'Código')

        self.assertEqual(filas, 1, 'solo el artículo de Bodega 2')

    def test_una_hoja_sin_datos_no_revienta(self):
        """Sin préstamos abiertos el archivo igual se genera, solo vacío."""
        hoja = self.hoja(reverse('reporte_prestamos'))
        self.assertEqual(hoja.max_row, self.fila_del_encabezado(hoja, 'Origen'))


class InventarioTecnicaTests(TestCase):
    """
    RF-12/RF-14: el listado de Bodega Técnica con su Excel.

    No existía ninguno: de esta bodega solo había cuatro números de
    valorización y la lista de lo prestado, así que no se podía sacar el
    inventario completo ni bajarlo como el de Bodega 1 y 2.
    """

    @classmethod
    def setUpTestData(cls):
        cls.btec = Bodega.objects.create(nombre='Bodega Técnica', tipo=Bodega.Tipo.TECNICA)
        cls.admin = Usuario.objects.create_user(
            username='admin_rt', password='clave-de-prueba', rol=Usuario.Rol.ADMINISTRADOR,
        )
        cls.taladro = Activo.objects.create(
            codigo_interno='SE-RT1', nombre_producto='Taladro', bodega=cls.btec, precio=900,
        )
        cls.bombillos = Activo.objects.create(
            codigo_interno='SE-RT2', nombre_producto='Bombillo LED', bodega=cls.btec,
            precio=25, es_consumible=True,
        )
        cls.sierra = Activo.objects.create(
            codigo_interno='SE-RT3', nombre_producto='Sierra vieja', bodega=cls.btec, precio=500,
        )
        dar_existencia(cls.taladro, 1, cls.admin)
        dar_existencia(cls.bombillos, 10, cls.admin)
        dar_existencia(cls.sierra, 1, cls.admin)
        dar_de_baja(cls.sierra, 1, cls.admin)

    def setUp(self):
        self.client.force_login(self.admin)

    def test_lista_lo_que_tiene_existencia(self):
        detalle, totales = reportes.inventario_tecnica()

        self.assertEqual(totales['productos'], 2, 'la sierra dada de baja no sale')
        self.assertEqual(totales['unidades'], 11)
        self.assertEqual(totales['valor'], 900 + 250)

    def test_se_pueden_incluir_los_dados_de_baja(self):
        detalle, totales = reportes.inventario_tecnica(solo_con_existencia=False)

        self.assertEqual(totales['productos'], 3)
        self.assertEqual(totales['valor'], 900 + 250, 'lo agotado no aporta valor')

    def test_cuenta_lo_que_esta_prestado(self):
        PrestamoActivo.objects.create(
            activo=self.bombillos, cantidad=4, solicitante='Byron', usuario=self.admin,
            estado_al_salir=Activo.Estado.BUEN_ESTADO,
        )

        _detalle, totales = reportes.inventario_tecnica()

        self.assertEqual(totales['afuera'], 4)
        self.assertEqual(totales['unidades'], 11, 'lo prestado sigue siendo de la bodega')

    def test_la_pantalla_responde(self):
        respuesta = self.client.get(reverse('reporte_tecnica'))

        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, 'Bombillo LED')

    def test_se_descarga_en_excel(self):
        respuesta = self.client.get(reverse('reporte_tecnica'), {'formato': 'excel'})

        self.assertEqual(respuesta.status_code, 200)
        self.assertIn('spreadsheetml', respuesta['Content-Type'])
        self.assertIn('bodega-tecnica', respuesta['Content-Disposition'])

    def test_contabilidad_tambien_lo_ve(self):
        """RF-04: consultar e imprimir lo pueden los tres roles."""
        contable = Usuario.objects.create_user(
            username='cont_rt', password='clave-de-prueba', rol=Usuario.Rol.CONTABILIDAD,
        )
        self.client.force_login(contable)

        self.assertEqual(self.client.get(reverse('reporte_tecnica')).status_code, 200)
