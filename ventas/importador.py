"""
Carga masiva de artículos desde Excel (RF-09).

Diseño en 2 pasos: (1) subir el archivo y elegir la hoja, (2) revisar el
mapeo de columnas sugerido (autodetectado por nombre de encabezado, siempre
editable) + previsualizar + confirmar.

A propósito NO se ofrece mapear "código interno": esa columna se ignora
aunque exista en el archivo (en el Excel real viene inconsistente) — el
código se genera solo con la misma regla que ya usa el catálogo
(Articulo.generar_codigo_interno, ver ventas/models.py).
"""

import re
from decimal import Decimal, InvalidOperation

import openpyxl

# (clave del campo, etiqueta en pantalla, obligatorio, palabras clave para autodetectar)
CAMPOS_IMPORTABLES = [
    ('nombre_producto', 'Producto', True, ['PRODUCTO']),
    ('bodega_col', 'Bodega (No. 1 / No. 2)', True, ['BODEGA']),
    ('marca', 'Marca', False, ['MARCA']),
    ('modelo', 'Modelo', False, ['MODELO']),
    ('capacidad', 'Capacidad', False, ['CAPACIDAD']),
    ('precio', 'Precio', False, ['PRECIO']),
    ('proveedor', 'Proveedor', False, ['PROVEEDOR']),
    ('numero_serie', 'Número de serie', False, ['SERIE', 'SERIAL']),
    ('stock_inicial', 'Stock / existencia inicial', False, ['TOTAL EXISTENCIA', 'EXISTENCIA MENSUAL', 'EXISTENCIA']),
]

MAX_FILAS_PREVIEW = 8
MAX_FILAS_ESCANEAR_ENCABEZADO = 12


def abrir_libro(ruta):
    return openpyxl.load_workbook(ruta, data_only=True, read_only=True)


def listar_hojas(ruta):
    libro = abrir_libro(ruta)
    try:
        return libro.sheetnames
    finally:
        libro.close()


def _fila_es_encabezado(valores):
    """Cuenta celdas de texto no vacías — la fila de encabezados real suele
    tener muchas más que cualquier fila de datos o título suelto."""
    return sum(1 for v in valores if isinstance(v, str) and v.strip())


def detectar_encabezados(ruta, hoja):
    """
    Escanea las primeras filas para encontrar la fila de encabezados (la que
    tiene más celdas de texto), y devuelve [(letra_columna, texto), ...]
    para las columnas que sí tienen encabezado.
    """
    libro = abrir_libro(ruta)
    try:
        ws = libro[hoja]
        mejor_fila, mejor_cuenta = 1, -1
        candidatas = []
        for i, fila in enumerate(ws.iter_rows(min_row=1, max_row=MAX_FILAS_ESCANEAR_ENCABEZADO, values_only=True), start=1):
            candidatas.append(fila)
            cuenta = _fila_es_encabezado(fila)
            if cuenta > mejor_cuenta:
                mejor_fila, mejor_cuenta = i, cuenta

        fila_valores = candidatas[mejor_fila - 1] if mejor_fila <= len(candidatas) else []
        columnas = []
        for idx, valor in enumerate(fila_valores):
            if isinstance(valor, str) and valor.strip():
                letra = openpyxl.utils.get_column_letter(idx + 1)
                columnas.append((letra, valor.strip()))
        return mejor_fila, columnas
    finally:
        libro.close()


def autodetectar_mapeo(columnas):
    """columnas: [(letra, texto), ...] -> {clave_campo: letra_columna_sugerida}"""
    mapeo = {}
    for clave, _etiqueta, _obligatorio, palabras in CAMPOS_IMPORTABLES:
        for letra, texto in columnas:
            texto_up = texto.upper()
            if any(p in texto_up for p in palabras):
                mapeo[clave] = letra
                break
    return mapeo


def _valor_texto(v):
    if v is None:
        return ''
    return str(v).strip()


def _valor_decimal(v):
    if v is None or v == '':
        return Decimal('0')
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    texto = str(v).strip().replace('Q', '').replace(',', '').replace(' ', '')
    try:
        return Decimal(texto) if texto else Decimal('0')
    except InvalidOperation:
        return Decimal('0')


def _valor_entero(v):
    if v is None or v == '':
        return 0
    try:
        return max(0, int(round(float(str(v).replace(',', '')))))
    except (ValueError, TypeError):
        return 0


def _numero_bodega(v):
    """'No. 2', 'No.2', 'NO.2', 2, '2' -> 2. None si no se reconoce."""
    if v is None:
        return None
    m = re.search(r'(\d+)', str(v))
    return int(m.group(1)) if m else None


def leer_filas(ruta, hoja, fila_encabezado, mapeo):
    """
    Generador de filas ya traducidas a los campos del sistema, a partir de
    la fila siguiente al encabezado hasta el final de la hoja. Filas sin
    "nombre_producto" se omiten (secciones en blanco entre bloques, comunes
    en estos archivos).
    """
    libro = abrir_libro(ruta)
    try:
        ws = libro[hoja]
        col_idx = {clave: openpyxl.utils.column_index_from_string(letra) - 1
                   for clave, letra in mapeo.items() if letra}

        for fila in ws.iter_rows(min_row=fila_encabezado + 1, values_only=True):
            def obtener(clave):
                idx = col_idx.get(clave)
                return fila[idx] if idx is not None and idx < len(fila) else None

            nombre = _valor_texto(obtener('nombre_producto'))
            if not nombre:
                continue

            yield {
                'nombre_producto': nombre,
                'bodega_num': _numero_bodega(obtener('bodega_col')),
                'marca': _valor_texto(obtener('marca')),
                'modelo': _valor_texto(obtener('modelo')),
                'capacidad': _valor_texto(obtener('capacidad')),
                'precio': _valor_decimal(obtener('precio')),
                'proveedor': _valor_texto(obtener('proveedor')),
                'numero_serie': _valor_texto(obtener('numero_serie')) or None,
                'stock_inicial': _valor_entero(obtener('stock_inicial')),
            }
    finally:
        libro.close()


def ejecutar_importacion(ruta, hoja, fila_encabezado, mapeo, usuario):
    """
    Crea o actualiza artículos a partir del archivo mapeado.

    "¿Ya existe?" se decide con el MISMO código que generaría el catálogo
    (SE-MODELO-CAPACIDAD) — no con lo que traiga la columna CODIGO INTERNO
    del Excel (esa se ignora a propósito, ver encabezado del archivo). Si
    ya existe, se actualiza; si no, se crea, y si trae stock inicial > 0 se
    registra como movimiento "Ajuste / Saldo inicial" (RF-09).

    Cada fila se guarda en su propio savepoint: si una fila falla (dato
    raro, duplicado, etc.) no se pierde el resto de la carga.
    """
    from django.db import IntegrityError, transaction

    from core.models import Bodega, Proveedor
    from .models import Articulo, MovimientoVenta

    bodegas_venta = {b.nombre: b for b in Bodega.objects.filter(tipo=Bodega.Tipo.VENTA)}

    resultado = {'creados': 0, 'actualizados': 0, 'omitidos': 0, 'errores': []}

    for i, fila in enumerate(leer_filas(ruta, hoja, fila_encabezado, mapeo), start=2):
        etiqueta_fila = f"Fila {i} ('{fila['nombre_producto']}')"
        bodega = bodegas_venta.get(f"Bodega {fila['bodega_num']}") if fila['bodega_num'] else None
        if not bodega:
            resultado['omitidos'] += 1
            resultado['errores'].append(f"{etiqueta_fila}: no se reconoce la bodega, se omitió.")
            continue

        try:
            with transaction.atomic():
                proveedor = None
                if fila['proveedor']:
                    proveedor, _ = Proveedor.objects.get_or_create(nombre=fila['proveedor'])

                # El mismo código que generaría el catálogo: así se sabe si
                # esta fila ya existe (actualizar) o es nueva (crear).
                codigo_prospecto = Articulo(modelo=fila['modelo'], capacidad=fila['capacidad']).generar_codigo_interno()
                existente = Articulo.objects.filter(codigo_interno=codigo_prospecto).first()

                if existente:
                    existente.nombre_producto = fila['nombre_producto']
                    existente.marca = fila['marca']
                    existente.modelo = fila['modelo']
                    existente.capacidad = fila['capacidad']
                    existente.bodega = bodega
                    existente.precio = fila['precio']
                    if proveedor:
                        existente.proveedor = proveedor
                    if fila['numero_serie']:
                        existente.numero_serie = fila['numero_serie']
                    existente.save()
                    resultado['actualizados'] += 1
                else:
                    articulo = Articulo.objects.create(
                        nombre_producto=fila['nombre_producto'],
                        marca=fila['marca'],
                        modelo=fila['modelo'],
                        capacidad=fila['capacidad'],
                        bodega=bodega,
                        precio=fila['precio'],
                        proveedor=proveedor,
                        numero_serie=fila['numero_serie'],
                    )
                    if fila['stock_inicial'] > 0:
                        MovimientoVenta.objects.create(
                            tipo_documento=MovimientoVenta.TipoDocumento.INGRESO,
                            tipo_transaccion=MovimientoVenta.TipoTransaccion.AJUSTE_INICIAL,
                            articulo=articulo,
                            cantidad=fila['stock_inicial'],
                            usuario=usuario,
                            observacion='Saldo inicial — carga masiva desde Excel.',
                        )
                    resultado['creados'] += 1
        except IntegrityError as e:
            resultado['omitidos'] += 1
            resultado['errores'].append(f"{etiqueta_fila}: {e}")

    return resultado
