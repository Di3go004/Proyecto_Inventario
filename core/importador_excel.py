"""
Utilidades genéricas para leer archivos .xlsx en la carga masiva (RF-09),
compartidas entre ventas/importador.py y tecnica/importador.py — ninguna
de estas funciones sabe nada de Articulo/Activo, solo de hojas y celdas.
"""

from decimal import Decimal, InvalidOperation

import openpyxl

MAX_FILAS_PREVIEW = 8
MAX_FILAS_ESCANEAR_ENCABEZADO = 12


def abrir_libro(ruta):
    return openpyxl.load_workbook(ruta, data_only=True, read_only=True)


def listar_hojas(ruta):
    libro = abrir_libro(ruta)
    try:
        return libro.sheetnames
    finally:
        libro.close()


def _fila_es_encabezado(valores):
    """Cuenta celdas de texto no vacías — la fila de encabezados real suele
    tener muchas más que cualquier fila de datos o título suelto."""
    return sum(1 for v in valores if isinstance(v, str) and v.strip())


def detectar_encabezados(ruta, hoja):
    """
    Escanea las primeras filas para encontrar la fila de encabezados (la que
    tiene más celdas de texto), y devuelve (numero_fila, [(letra, texto), ...])
    para las columnas que sí tienen encabezado.
    """
    libro = abrir_libro(ruta)
    try:
        ws = libro[hoja]
        mejor_fila, mejor_cuenta = 1, -1
        candidatas = []
        for i, fila in enumerate(ws.iter_rows(min_row=1, max_row=MAX_FILAS_ESCANEAR_ENCABEZADO, values_only=True), start=1):
            candidatas.append(fila)
            cuenta = _fila_es_encabezado(fila)
            if cuenta > mejor_cuenta:
                mejor_fila, mejor_cuenta = i, cuenta

        fila_valores = candidatas[mejor_fila - 1] if mejor_fila <= len(candidatas) else []
        columnas = []
        for idx, valor in enumerate(fila_valores):
            if isinstance(valor, str) and valor.strip():
                letra = openpyxl.utils.get_column_letter(idx + 1)
                columnas.append((letra, valor.strip()))
        return mejor_fila, columnas
    finally:
        libro.close()


def autodetectar_mapeo(columnas, campos_importables):
    """
    columnas: [(letra, texto), ...]; campos_importables: misma forma que
    CAMPOS_IMPORTABLES de cada módulo -> {clave_campo: letra_columna}.

    Manda el orden de las palabras clave, no el de las columnas: se busca la
    primera palabra en toda la hoja, y solo si no aparece se prueba la
    siguiente. Antes ganaba la columna que estuviera más a la izquierda, y en
    los Excel reales eso elegía mal: "EXISTENCIA INICIO SEMANA" viene mucho
    antes que "TOTAL EXISTENCIA MENSUAL", así que la carga masiva proponía el
    saldo del arranque de la semana 1 en vez del total del mes.
    """
    mapeo = {}
    for clave, _etiqueta, _obligatorio, palabras in campos_importables:
        for palabra in palabras:
            letra = next((l for l, texto in columnas if palabra in texto.upper()), None)
            if letra:
                mapeo[clave] = letra
                break
    return mapeo


def valor_texto(v):
    if v is None:
        return ''
    return str(v).strip()


def valor_decimal(v):
    if v is None or v == '':
        return Decimal('0')
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    texto = str(v).strip().replace('Q', '').replace(',', '').replace(' ', '')
    try:
        return Decimal(texto) if texto else Decimal('0')
    except InvalidOperation:
        return Decimal('0')


def valor_entero(v):
    if v is None or v == '':
        return 0
    try:
        return max(0, int(round(float(str(v).replace(',', '')))))
    except (ValueError, TypeError):
        return 0


def columnas_a_indices(mapeo):
    """{'clave': 'H'} -> {'clave': 7} (índice base-0 dentro de la fila)."""
    return {clave: openpyxl.utils.column_index_from_string(letra) - 1 for clave, letra in mapeo.items() if letra}
