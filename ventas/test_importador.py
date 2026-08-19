"""
Pruebas de la carga masiva desde Excel (RF-09).

Se construye un .xlsx en memoria que imita la estructura real de
FO-SE-053 (fila de encabezados abajo de un título, columna "No. BODEGA"
con el formato "No. 1" / "No. 2") para comprobar el camino completo:
detectar encabezados → mapear columnas → importar → stock inicial.
"""

import tempfile

import openpyxl
from django.test import TestCase

from core.models import Bodega, Proveedor
from usuarios.models import Usuario
from ventas import importador
from ventas.models import Articulo, MovimientoVenta


def crear_excel_de_prueba(filas):
    """Genera un .xlsx con la misma forma que el archivo real de la empresa."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = 'EQUIPO ENERO 2025'

    # El archivo real trae un título y una fila en blanco antes de los encabezados.
    hoja['A1'] = 'SOLUCIONES EXACTAS, S.A.'
    hoja['A3'] = 'LISTADO INVENTARIO BODEGAS'
    encabezados = ['No.', 'PRODUCTO', 'No. BODEGA', 'MARCA', 'MODELO', 'CAPACIDAD',
                   'PRECIO', 'PROVEEDOR', 'TOTAL EXISTENCIA MENSUAL']
    hoja.append([])
    for columna, texto in enumerate(encabezados, start=1):
        hoja.cell(row=4, column=columna, value=texto)
    for indice, fila in enumerate(filas, start=5):
        for columna, valor in enumerate(fila, start=1):
            hoja.cell(row=indice, column=columna, value=valor)

    temporal = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    libro.save(temporal.name)
    temporal.close()
    return temporal.name


class CargaMasivaTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.bodega1 = Bodega.objects.create(nombre='Bodega 1', tipo=Bodega.Tipo.VENTA)
        cls.bodega2 = Bodega.objects.create(nombre='Bodega 2', tipo=Bodega.Tipo.VENTA)
        cls.admin = Usuario.objects.create_user(
            username='admin_import', password='clave-de-prueba', rol=Usuario.Rol.ADMINISTRADOR,
        )

    def importar(self, filas):
        ruta = crear_excel_de_prueba(filas)
        hoja = 'EQUIPO ENERO 2025'
        fila_encabezado, columnas = importador.detectar_encabezados(ruta, hoja)
        mapeo = importador.autodetectar_mapeo(columnas)
        return importador.ejecutar_importacion(ruta, hoja, fila_encabezado, mapeo, self.admin), mapeo

    def test_detecta_los_encabezados_aunque_no_esten_en_la_primera_fila(self):
        ruta = crear_excel_de_prueba([[1, 'Bascula', 'No. 1', 'Marca', 'M-1', '10kg', 100, 'Prov', 5]])
        fila, columnas = importador.detectar_encabezados(ruta, 'EQUIPO ENERO 2025')

        self.assertEqual(fila, 4, 'los encabezados del archivo real estan en la fila 4')
        self.assertIn(('B', 'PRODUCTO'), columnas)
        self.assertIn(('C', 'No. BODEGA'), columnas)

    def test_autodetecta_el_mapeo_de_columnas(self):
        ruta = crear_excel_de_prueba([[1, 'Bascula', 'No. 1', 'Marca', 'M-1', '10kg', 100, 'Prov', 5]])
        _, columnas = importador.detectar_encabezados(ruta, 'EQUIPO ENERO 2025')
        mapeo = importador.autodetectar_mapeo(columnas)

        self.assertEqual(mapeo['nombre_producto'], 'B')
        self.assertEqual(mapeo['bodega_col'], 'C')
        self.assertEqual(mapeo['precio'], 'G')

    def test_importa_creando_articulos_con_su_stock_inicial(self):
        resultado, _ = self.importar([
            [1, 'Bascula de plataforma', 'No. 1', 'Brecknell', 'BP-300', '300kg', 1500, 'SIZEIN', 7],
            [2, 'Adaptador 9V', 'No. 2', 'Keerda', 'DZ012', '9V', 395, 'LOCOSC', 3],
        ])

        self.assertEqual(resultado['creados'], 2)
        self.assertEqual(resultado['omitidos'], 0)

        bascula = Articulo.objects.get(nombre_producto='Bascula de plataforma')
        self.assertEqual(bascula.bodega, self.bodega1)
        self.assertEqual(bascula.stock_actual, 7)
        self.assertEqual(str(bascula.precio), '1500.00')

        adaptador = Articulo.objects.get(nombre_producto='Adaptador 9V')
        self.assertEqual(adaptador.bodega, self.bodega2)
        self.assertEqual(adaptador.stock_actual, 3)

    def test_el_stock_inicial_queda_como_movimiento_de_ajuste(self):
        """Así el stock importado también tiene su rastro en el historial."""
        self.importar([[1, 'Bascula', 'No. 1', 'Brecknell', 'BP-300', '300kg', 1500, 'SIZEIN', 7]])

        movimiento = MovimientoVenta.objects.get()
        self.assertEqual(movimiento.tipo_transaccion, MovimientoVenta.TipoTransaccion.AJUSTE_INICIAL)
        self.assertEqual(movimiento.tipo_documento, MovimientoVenta.TipoDocumento.INGRESO)
        self.assertEqual(movimiento.cantidad, 7)

    def test_ignora_el_codigo_del_archivo_y_aplica_el_estandar_de_la_empresa(self):
        """RF-09: el código interno del Excel viene inconsistente, así que se
        regenera como SE-MODELO-capacidad."""
        self.importar([[1, 'Adaptador', 'No. 2', 'Keerda', 'DZ012ELL', '9V', 395, 'LOCOSC', 1]])

        articulo = Articulo.objects.get(nombre_producto='Adaptador')
        self.assertEqual(articulo.codigo_interno, 'SE-DZ012ELL-9V')

    def test_reimportar_la_misma_hoja_actualiza_en_vez_de_duplicar(self):
        filas = [[1, 'Bascula', 'No. 1', 'Brecknell', 'BP-300', '300kg', 1500, 'SIZEIN', 7]]
        self.importar(filas)

        filas[0][6] = 1800  # cambio de precio en el archivo
        resultado, _ = self.importar(filas)

        self.assertEqual(resultado['creados'], 0)
        self.assertEqual(resultado['actualizados'], 1)
        self.assertEqual(Articulo.objects.count(), 1)
        self.assertEqual(str(Articulo.objects.get().precio), '1800.00')

    def test_omite_las_filas_sin_bodega_reconocible_y_avisa(self):
        """En el archivo real hay filas con 'N/A' en la columna de bodega."""
        resultado, _ = self.importar([
            [1, 'Bascula buena', 'No. 1', 'Brecknell', 'BP-300', '300kg', 1500, 'SIZEIN', 7],
            [2, 'Balanza sin bodega', 'N/A', 'Marca', 'M-2', '1kg', 100, 'Prov', 2],
        ])

        self.assertEqual(resultado['creados'], 1)
        self.assertEqual(resultado['omitidos'], 1)
        self.assertEqual(len(resultado['errores']), 1)
        self.assertIn('no se reconoce la bodega', resultado['errores'][0])
        self.assertFalse(Articulo.objects.filter(nombre_producto='Balanza sin bodega').exists())

    def test_reutiliza_los_proveedores_en_vez_de_duplicarlos(self):
        self.importar([
            [1, 'Producto A', 'No. 1', 'Marca', 'M-A', '1kg', 100, 'SIZEIN', 1],
            [2, 'Producto B', 'No. 1', 'Marca', 'M-B', '2kg', 200, 'SIZEIN', 1],
        ])
        self.assertEqual(Proveedor.objects.filter(nombre='SIZEIN').count(), 1)

    def test_las_filas_en_blanco_no_generan_articulos(self):
        resultado, _ = self.importar([
            [1, 'Producto real', 'No. 1', 'Marca', 'M-1', '1kg', 100, 'Prov', 1],
            [None, None, None, None, None, None, None, None, None],
            [3, '', 'No. 1', 'Marca', 'M-3', '3kg', 300, 'Prov', 1],
        ])
        self.assertEqual(resultado['creados'], 1)
        self.assertEqual(Articulo.objects.count(), 1)
