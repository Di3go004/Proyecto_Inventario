"""
Cómo se escribe "no tiene número de serie".

En la empresa lo abrevian **S/S**, así que es lo que aparece en pantalla y en
los reportes cuando el campo está vacío, en vez de una raya.

Es una forma de escribirlo, no un dato: en la base el campo sigue vacío. Si se
guardara el texto, un artículo sin serial sería indistinguible de uno cuyo
serial fuera literalmente "S/S", y las búsquedas por serial se romperían.
"""

import io

import openpyxl
from django.test import TestCase
from django.urls import reverse

from core.models import Bodega
from usuarios.models import Usuario
from ventas.models import SIN_SERIAL, Articulo, limpiar_serial


class BaseSerial(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.bodega = Bodega.objects.create(nombre='Bodega 1', tipo=Bodega.Tipo.VENTA)
        cls.admin = Usuario.objects.create_user(
            username='admin_serial', password='clave-de-prueba', rol=Usuario.Rol.ADMINISTRADOR,
        )
        cls.con_serial = Articulo.objects.create(
            nombre_producto='BASCULA CON SERIAL', modelo='B-1', bodega=cls.bodega,
            numero_serie='AB-12345', precio=100,
        )
        cls.sin_serial = Articulo.objects.create(
            nombre_producto='BASCULA SIN SERIAL', modelo='B-2', bodega=cls.bodega,
            precio=200,
        )

    def setUp(self):
        self.client.force_login(self.admin)


class PropiedadTests(BaseSerial):
    def test_el_que_tiene_serial_lo_muestra(self):
        self.assertEqual(self.con_serial.serial, 'AB-12345')

    def test_el_que_no_tiene_muestra_ss(self):
        self.assertEqual(self.sin_serial.serial, SIN_SERIAL)
        self.assertEqual(self.sin_serial.serial, 'S/S')

    def test_en_la_base_el_campo_sigue_vacio(self):
        """
        Guardar el texto haría imposible distinguir "no tiene" de un serial
        que se llamara así, y rompería buscar por serial.

        El campo queda en NULL y no en cadena vacía porque es único: con dos
        artículos sin serial, dos cadenas vacías chocarían entre sí. SQL sí
        permite varios NULL en un índice único.
        """
        self.sin_serial.refresh_from_db()
        self.assertIsNone(self.sin_serial.numero_serie)
        self.assertFalse(
            Articulo.objects.filter(numero_serie=SIN_SERIAL).exists(),
            'el "S/S" no se guarda: solo se muestra',
        )

    def test_pueden_existir_varios_sin_serial(self):
        """Es el motivo de que el campo sea NULL y no cadena vacía."""
        Articulo.objects.create(
            nombre_producto='OTRA SIN SERIAL', modelo='B-3', bodega=self.bodega,
        )

        sin_serial = Articulo.objects.filter(numero_serie__isnull=True)

        self.assertGreaterEqual(sin_serial.count(), 2)
        self.assertTrue(all(a.serial == SIN_SERIAL for a in sin_serial))


class PantallasTests(BaseSerial):
    def test_el_catalogo_lo_muestra(self):
        respuesta = self.client.get(reverse('catalogo_articulos'))

        self.assertContains(respuesta, 'S/S')
        self.assertContains(respuesta, 'AB-12345')

    def test_la_ficha_lo_muestra(self):
        respuesta = self.client.get(reverse('articulo_detalle', args=[self.sin_serial.pk]))

        self.assertContains(respuesta, 'S/S')

    def test_la_ficha_del_que_si_tiene_no_dice_ss(self):
        respuesta = self.client.get(reverse('articulo_detalle', args=[self.con_serial.pk]))

        self.assertContains(respuesta, 'AB-12345')
        self.assertNotContains(respuesta, 'S/S')

    def test_ya_no_sale_la_raya(self):
        """Antes se pintaba "—", que no es como lo escriben en la empresa."""
        respuesta = self.client.get(reverse('articulo_detalle', args=[self.sin_serial.pk]))
        cuerpo = respuesta.content.decode()

        # La raya sigue usándose en otros campos vacíos; lo que importa es que
        # la fila del serial ya no la lleve.
        inicio = cuerpo.index('Número de serial')
        self.assertIn('S/S', cuerpo[inicio:inicio + 200])


class FiltroDePlantillaTests(TestCase):
    """
    La vista previa de la carga masiva trabaja con filas del Excel todavía sin
    guardar, no con artículos, así que no puede usar la propiedad del modelo.
    """

    def test_el_filtro_hace_lo_mismo_que_la_propiedad(self):
        from ventas.templatetags.catalogo_extras import serial

        self.assertEqual(serial('AB-12345'), 'AB-12345')
        self.assertEqual(serial(''), SIN_SERIAL)
        self.assertEqual(serial(None), SIN_SERIAL)


class ReporteDeExistenciasTests(BaseSerial):
    """
    El reporte que lista todos los artículos ahora trae el serial, que es
    donde hacía falta para poder cruzarlo con el inventario físico.
    """

    def hoja(self):
        respuesta = self.client.get(reverse('reporte_existencias'), {'formato': 'excel'})
        self.assertEqual(respuesta.status_code, 200)
        return openpyxl.load_workbook(io.BytesIO(respuesta.content)).active

    def fila_del_encabezado(self, hoja):
        for numero in range(1, 15):
            if hoja.cell(row=numero, column=1).value == 'Código':
                return numero
        self.fail('no se encontró el encabezado de la hoja')

    def columnas(self, hoja):
        fila = self.fila_del_encabezado(hoja)
        titulos = {}
        for columna in range(1, 20):
            valor = hoja.cell(row=fila, column=columna).value
            if valor:
                titulos[valor] = columna
        return fila, titulos

    def test_la_pantalla_muestra_el_serial(self):
        respuesta = self.client.get(reverse('reporte_existencias'))

        self.assertContains(respuesta, 'AB-12345')
        self.assertContains(respuesta, 'S/S')

    def test_el_excel_trae_la_columna(self):
        hoja = self.hoja()
        _fila, titulos = self.columnas(hoja)

        self.assertIn('N.º de serial', titulos)

    def test_el_excel_escribe_ss_donde_no_hay(self):
        hoja = self.hoja()
        fila_enc, titulos = self.columnas(hoja)
        columna = titulos['N.º de serial']

        seriales = [
            hoja.cell(row=f, column=columna).value
            for f in range(fila_enc + 1, fila_enc + 3)
        ]

        self.assertIn('AB-12345', seriales)
        self.assertIn(SIN_SERIAL, seriales)

    def test_el_formato_de_moneda_sigue_en_las_columnas_de_dinero(self):
        """
        Regresión: al insertar la columna del serial se corren todas las que
        van después. Los formatos de moneda se indican por número de columna,
        así que si no se corren también terminan pintando la existencia o el
        nivel como si fueran quetzales.
        """
        from core import exportar

        hoja = self.hoja()
        fila_enc, titulos = self.columnas(hoja)
        primera_de_datos = fila_enc + 1

        for titulo in ('Precio unitario', 'Valor total'):
            with self.subTest(columna=titulo):
                celda = hoja.cell(row=primera_de_datos, column=titulos[titulo])
                self.assertEqual(celda.number_format, exportar.FORMATO_MONEDA)

        for titulo in ('Existencia', 'Nivel'):
            with self.subTest(columna=titulo):
                celda = hoja.cell(row=primera_de_datos, column=titulos[titulo])
                self.assertNotEqual(
                    celda.number_format, exportar.FORMATO_MONEDA,
                    f'"{titulo}" no es dinero y no debe salir con formato de quetzales',
                )


class EscribirloAManoTests(BaseSerial):
    """
    "S/S" está a la vista en todas las pantallas y en el manual, así que
    alguien lo va a escribir en el campo. Se toma como lo que significa —no
    tiene serial— y no como un serial que se llama así.

    Sin esto el segundo artículo al que se lo escribieran chocaría contra el
    índice único con "ya existe un artículo con ese número de serie", que no
    dice nada a quien solo quiso anotar que la báscula no traía placa.
    """

    def guardar(self, serial):
        respuesta = self.client.post(reverse('articulo_nuevo'), {
            'codigo_interno': '', 'numero_serie': serial,
            'nombre_producto': f'BASCULA {serial or "VACIA"}', 'marca': '',
            'modelo': f'M-{abs(hash(serial)) % 10000}', 'capacidad': '',
            'bodega': self.bodega.pk, 'categoria': '', 'proveedor': '',
            'precio': '100', 'imagen_url': '',
            'stock_optimo': 0, 'stock_alerta': 0, 'stock_critico': 0,
            'activo': 'on',
        })
        self.assertEqual(respuesta.status_code, 302, 'el formulario no guardó')
        return Articulo.objects.latest('id')

    def test_escribir_ss_es_decir_que_no_tiene(self):
        articulo = self.guardar('S/S')

        self.assertIsNone(articulo.numero_serie)
        self.assertEqual(articulo.serial, SIN_SERIAL)

    def test_dos_articulos_pueden_llevar_ss_escrito(self):
        """Es el choque contra el índice único que esto viene a evitar."""
        primero = self.guardar('S/S')
        segundo = self.guardar('s/s')

        self.assertIsNone(primero.numero_serie)
        self.assertIsNone(segundo.numero_serie)
        self.assertNotEqual(primero.pk, segundo.pk)

    def test_un_serial_de_verdad_se_guarda_igual(self):
        articulo = self.guardar('AB-999')

        self.assertEqual(articulo.numero_serie, 'AB-999')

    def test_le_quita_los_espacios_de_los_lados(self):
        articulo = self.guardar('  AB-777  ')

        self.assertEqual(articulo.numero_serie, 'AB-777')


class LimpiarSerialTests(TestCase):
    """Las formas en que llega escrito "no tiene", del Excel y del teclado."""

    def test_las_que_significan_que_no_tiene(self):
        for escrito in ('', '   ', 'S/S', 's/s', 'S / S', 'sin serie', 'SIN SERIAL',
                        '-', '-----', '.', None):
            with self.subTest(escrito=escrito):
                self.assertIsNone(limpiar_serial(escrito))

    def test_las_que_si_son_un_serial(self):
        for escrito in ('AB-12345', '0', 'S/S-4471', 'SN-1'):
            with self.subTest(escrito=escrito):
                self.assertEqual(limpiar_serial(escrito), escrito)
