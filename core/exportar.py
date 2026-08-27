"""
Exportación de reportes a Excel (RF-14).

Se usa openpyxl, que ya estaba como dependencia para leer los archivos de la
carga masiva (RF-09) — no agrega nada nuevo al proyecto.

Los archivos salen listos para trabajar: encabezado fijo al hacer scroll,
filtros activados, anchos ajustados y los montos con formato de moneda. Un
reporte que hay que reformatear a mano cada vez termina no usándose.
"""

import io
from datetime import date, datetime

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

EMPRESA = 'SOLUCIONES EXACTAS, S.A.'

AZUL = '202199'
BLANCO = 'FFFFFF'
FORMATO_MONEDA = '"Q" #,##0.00'
FORMATO_FECHA = 'DD/MM/YYYY HH:MM'

# Tres filas de titulo, una en blanco, y la tabla arranca aqui.
FILA_ENCABEZADO = 5


def _ancho_sugerido(columna):
    """Ancho aproximado según el contenido más largo, con topes razonables."""
    mayor = 0
    for celda in columna:
        if celda.value is not None:
            mayor = max(mayor, len(str(celda.value)))
    return min(max(mayor + 3, 10), 55)


def libro(titulo, encabezados, filas, subtitulo='', formatos=None, nombre_hoja=None):
    """
    Arma un .xlsx de una sola hoja y devuelve sus bytes.

    `formatos` mapea el índice de columna (0-based) a un formato de número de
    Excel, para que los montos salgan como moneda y las fechas como fecha en
    vez de como texto.
    """
    formatos = formatos or {}

    documento = Workbook()
    hoja = documento.active
    # Excel no admite : \ / ? * [ ] en el nombre de la hoja, ni más de 31.
    hoja.title = (nombre_hoja or titulo)[:31].replace('/', '-').replace(':', '-')

    # Se escribe por número de fila y no con append(): appendear una fila
    # vacía adelanta el cursor interno de openpyxl pero no crea celdas, así
    # que max_row se queda atrás y todo lo que dependa de él (el formato del
    # encabezado, el panel fijo, el rango del filtro) termina una fila corrido.
    hoja.cell(row=1, column=1, value=EMPRESA).font = Font(bold=True, size=13)
    hoja.cell(row=2, column=1, value=titulo).font = Font(bold=True, size=11)
    pie = f'Generado el {timezone.localtime():%d/%m/%Y %H:%M}'
    hoja.cell(
        row=3, column=1, value=f'{subtitulo} · {pie}' if subtitulo else pie,
    ).font = Font(italic=True, size=9, color='666666')
    # La fila 4 se deja en blanco, para separar el título de la tabla.

    for columna, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=FILA_ENCABEZADO, column=columna, value=texto)
        celda.font = Font(bold=True, color=BLANCO)
        celda.fill = PatternFill('solid', fgColor=AZUL)
        celda.alignment = Alignment(vertical='center', wrap_text=True)

    for desplazamiento, fila in enumerate(filas):
        for columna, valor in enumerate(fila, start=1):
            hoja.cell(row=FILA_ENCABEZADO + 1 + desplazamiento, column=columna, value=valor)

    for indice, formato in formatos.items():
        letra = get_column_letter(indice + 1)
        for celda in hoja[letra][FILA_ENCABEZADO:]:
            celda.number_format = formato

    for columna in hoja.columns:
        letra = get_column_letter(columna[0].column)
        hoja.column_dimensions[letra].width = _ancho_sugerido(columna)

    if filas:
        # El encabezado queda fijo al hacer scroll y con filtros puestos.
        hoja.freeze_panes = hoja.cell(row=FILA_ENCABEZADO + 1, column=1)
        hoja.auto_filter.ref = (
            f'A{FILA_ENCABEZADO}:'
            f'{get_column_letter(len(encabezados))}{FILA_ENCABEZADO + len(filas)}'
        )

    memoria = io.BytesIO()
    documento.save(memoria)
    return memoria.getvalue()


def valor_para_excel(valor):
    """
    Deja los valores en un tipo que Excel entienda como número o fecha, para
    que se puedan sumar y ordenar. Un reporte donde los montos llegan como
    texto obliga a reescribirlo entero antes de usarlo.
    """
    if isinstance(valor, datetime):
        return timezone.localtime(valor).replace(tzinfo=None)
    if isinstance(valor, date):
        return valor
    return valor


def nombre_de_archivo(base):
    return f'{base}-{timezone.localtime():%Y-%m-%d}.xlsx'
