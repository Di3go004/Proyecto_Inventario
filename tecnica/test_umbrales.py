"""
Umbrales de reposición en Bodega Técnica (RF-11).

Funcionan igual que en Bodega 1 y 2 —mismos campos, mismos valores por
defecto, mismas restricciones de base— con una diferencia que es la decisión
de fondo: **se comparan contra la existencia, no contra lo disponible.**

Si miraran lo disponible, prestar tres de cuatro taladros dispararía "hay que
comprar taladros", cuando en realidad van a volver. Es la misma razón por la
que un préstamo no mueve la existencia: la herramienta sigue siendo de la
bodega.
"""

from django.db import IntegrityError, transaction
from django.test import TestCase
from django.urls import reverse

from core import reportes
from core.models import Bodega, UMBRALES_EN_ORDEN
from tecnica.models import Activo, MovimientoActivo, PrestamoActivo
from usuarios.models import Usuario


class BaseUmbrales(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.bodega = Bodega.objects.create(nombre='Bodega Técnica', tipo=Bodega.Tipo.TECNICA)
        cls.admin = Usuario.objects.create_user(
            username='admin_umbrales', password='clave-de-prueba', rol=Usuario.Rol.ADMINISTRADOR,
        )

    def setUp(self):
        self.client.force_login(self.admin)

    def activo(self, codigo='SE-TE700', existencia=0, **extra):
        activo = Activo.objects.create(
            codigo_interno=codigo, nombre_producto=f'Cosa {codigo}',
            bodega=self.bodega, precio=10, **extra,
        )
        if existencia:
            MovimientoActivo.objects.create(
                tipo=MovimientoActivo.Tipo.INGRESO, activo=activo,
                cantidad=existencia, usuario=self.admin, folio='A-1',
            )
            activo.refresh_from_db()
        return activo


class LosCamposTests(BaseUmbrales):
    def test_entran_con_los_mismos_valores_que_bodega_1_y_2(self):
        activo = self.activo()

        self.assertEqual(activo.stock_critico, 2)
        self.assertEqual(activo.stock_alerta, 5)
        self.assertEqual(activo.stock_optimo, 20)

    def test_son_los_mismos_que_los_de_un_articulo(self):
        """Que no se desincronicen: se pidió que funcionen igual."""
        from ventas.models import Articulo

        for campo in ('stock_critico', 'stock_alerta', 'stock_optimo'):
            with self.subTest(campo=campo):
                self.assertEqual(
                    Activo._meta.get_field(campo).default,
                    Articulo._meta.get_field(campo).default,
                )

    def test_se_pueden_configurar_desde_el_formulario(self):
        activo = self.activo()

        self.client.post(reverse('activo_editar', args=[activo.pk]), {
            'codigo_interno': activo.codigo_interno, 'nombre_producto': activo.nombre_producto,
            'marca': '', 'modelo': '', 'bodega': self.bodega.pk, 'categoria': '',
            'proveedor': '', 'precio': '10', 'estado': Activo.Estado.BUEN_ESTADO,
            'imagen_url': '', 'stock_critico': 10, 'stock_alerta': 25, 'stock_optimo': 100,
        })

        activo.refresh_from_db()
        self.assertEqual(activo.stock_critico, 10)
        self.assertEqual(activo.stock_alerta, 25)
        self.assertEqual(activo.stock_optimo, 100)

    def test_la_pantalla_los_ofrece(self):
        respuesta = self.client.get(reverse('activo_nuevo'))

        self.assertContains(respuesta, 'id_stock_critico')
        self.assertContains(respuesta, 'id_stock_alerta')
        self.assertContains(respuesta, 'id_stock_optimo')


class ElOrdenLoObligaLaBaseTests(BaseUmbrales):
    """
    Igual que en Articulo: lo garantiza la restricción y no solo el
    formulario, para que no entre por otra puerta (admin, consola, carga
    masiva) una combinación que no tiene sentido.
    """

    def test_no_acepta_critico_mayor_que_alerta(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            Activo.objects.create(
                codigo_interno='SE-MALO', nombre_producto='Malo', bodega=self.bodega,
                stock_critico=9, stock_alerta=3, stock_optimo=20,
            )

    def test_no_acepta_alerta_mayor_que_optimo(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            Activo.objects.create(
                codigo_interno='SE-MALO2', nombre_producto='Malo', bodega=self.bodega,
                stock_critico=1, stock_alerta=30, stock_optimo=20,
            )

    def test_el_formulario_lo_explica_en_castellano(self):
        activo = self.activo()

        respuesta = self.client.post(reverse('activo_editar', args=[activo.pk]), {
            'codigo_interno': activo.codigo_interno, 'nombre_producto': activo.nombre_producto,
            'marca': '', 'modelo': '', 'bodega': self.bodega.pk, 'categoria': '',
            'proveedor': '', 'precio': '10', 'estado': Activo.Estado.BUEN_ESTADO,
            'imagen_url': '', 'stock_critico': 50, 'stock_alerta': 5, 'stock_optimo': 20,
        })

        self.assertEqual(respuesta.status_code, 200, 'debió volver al formulario')
        self.assertContains(respuesta, UMBRALES_EN_ORDEN)

    def test_el_mensaje_es_el_mismo_en_las_dos_bodegas(self):
        self.assertIn('crítico', UMBRALES_EN_ORDEN)
        self.assertIn('óptimo', UMBRALES_EN_ORDEN)


class ElNivelTests(BaseUmbrales):
    def test_en_cero_es_critico(self):
        self.assertEqual(self.activo(existencia=0).nivel_alerta, 'critico')

    def test_en_el_umbral_critico_todavia_es_critico(self):
        self.assertEqual(self.activo(existencia=2).nivel_alerta, 'critico')

    def test_entre_critico_y_alerta_es_alerta(self):
        self.assertEqual(self.activo(existencia=4).nivel_alerta, 'alerta')

    def test_entre_alerta_y_optimo_es_normal(self):
        self.assertEqual(self.activo(existencia=10).nivel_alerta, 'normal')

    def test_en_el_optimo_es_optimo(self):
        self.assertEqual(self.activo(existencia=20).nivel_alerta, 'optimo')

    def test_calcula_igual_que_un_articulo(self):
        """Se pidió que funcionen igual, así que se comparan uno contra otro."""
        from ventas.models import Articulo

        venta = Bodega.objects.create(nombre='Bodega 1', tipo=Bodega.Tipo.VENTA)
        for cuantos in (0, 2, 3, 5, 12, 20, 50):
            with self.subTest(existencia=cuantos):
                articulo = Articulo(
                    nombre_producto='x', modelo=f'm{cuantos}', bodega=venta,
                    stock_actual=cuantos,
                )
                activo = self.activo(codigo=f'SE-CMP{cuantos}', existencia=cuantos)
                self.assertEqual(activo.nivel_alerta, articulo.nivel_alerta)


class SeMideContraLaExistenciaTests(BaseUmbrales):
    """
    La decisión de fondo. Prestar no es gastar: la herramienta va a volver,
    así que un préstamo no debe disparar la alerta de reposición.
    """

    def test_prestar_casi_todo_no_dispara_la_alerta(self):
        activo = self.activo(existencia=10)

        PrestamoActivo.objects.create(
            activo=activo, cantidad=9, solicitante='Ivan Leiva',
            usuario=self.admin, estado_al_salir=Activo.Estado.BUEN_ESTADO,
        )

        activo.refresh_from_db()
        self.assertEqual(activo.disponibles, 1, 'queda una sola libre')
        self.assertEqual(activo.existencia, 10, 'pero la bodega sigue teniendo diez')
        self.assertEqual(activo.nivel_alerta, 'normal', 'prestar no es gastar')

    def test_dar_de_baja_si_la_dispara(self):
        """Lo que sí baja la existencia de verdad."""
        activo = self.activo(existencia=10)

        MovimientoActivo.objects.create(
            tipo=MovimientoActivo.Tipo.BAJA, activo=activo, cantidad=8,
            usuario=self.admin, motivo=MovimientoActivo.Motivo.DANADO,
        )

        activo.refresh_from_db()
        self.assertEqual(activo.existencia, 2)
        self.assertEqual(activo.nivel_alerta, 'critico')


class EnLasPantallasTests(BaseUmbrales):
    def test_el_catalogo_muestra_la_columna_nivel(self):
        self.activo(existencia=0)

        respuesta = self.client.get(reverse('catalogo_activos'))

        self.assertContains(respuesta, 'Nivel')
        self.assertContains(respuesta, '<span class="chip chip-critical">Crítico</span>', html=True)

    def test_la_ficha_muestra_el_nivel_y_los_umbrales(self):
        activo = self.activo(existencia=3)

        respuesta = self.client.get(reverse('activo_detalle', args=[activo.pk]))

        self.assertContains(respuesta, 'Nivel de reposición')
        self.assertContains(respuesta, '<span class="chip chip-warn">Alerta</span>', html=True)
        self.assertContains(respuesta, '2 / 5 / 20')


class EnElReporteDeAlertasTests(BaseUmbrales):
    def test_la_funcion_solo_trae_lo_que_esta_bajo_el_umbral(self):
        self.activo(codigo='SE-BAJO', existencia=1)
        self.activo(codigo='SE-BIEN', existencia=50)

        codigos = [a.codigo_interno for a in reportes.alertas_tecnica()]

        self.assertIn('SE-BAJO', codigos)
        self.assertNotIn('SE-BIEN', codigos)

    def test_lo_mas_urgente_va_primero(self):
        self.activo(codigo='SE-ALERTA', existencia=4)
        self.activo(codigo='SE-CERO', existencia=0)

        codigos = [a.codigo_interno for a in reportes.alertas_tecnica()]

        self.assertEqual(codigos[0], 'SE-CERO')

    def test_la_pantalla_lo_muestra_en_su_propia_seccion(self):
        self.activo(codigo='SE-TECNICO', existencia=1)

        respuesta = self.client.get(reverse('reporte_alertas'))

        self.assertContains(respuesta, 'Bodega Técnica')
        self.assertContains(respuesta, 'SE-TECNICO')

    def test_no_se_mezcla_con_bodega_1_y_2(self):
        """Van en tablas distintas: un activo tiene otras columnas."""
        self.activo(codigo='SE-TECNICO', existencia=1)

        respuesta = self.client.get(reverse('reporte_alertas'))

        self.assertIn('activos', respuesta.context)
        self.assertNotIn(
            'SE-TECNICO',
            [a.codigo_interno for a in respuesta.context['pagina']],
            'el activo no debe salir en la tabla de Bodega 1 y 2',
        )

    def test_el_excel_trae_una_hoja_para_bodega_tecnica(self):
        import io

        import openpyxl

        self.activo(codigo='SE-TECNICO', existencia=1)

        respuesta = self.client.get(reverse('reporte_alertas'), {'formato': 'excel'})

        libro = openpyxl.load_workbook(io.BytesIO(respuesta.content))
        self.assertEqual(len(libro.sheetnames), 2, f'hojas: {libro.sheetnames}')
        hoja = libro[[h for h in libro.sheetnames if 'Tecnica' in h][0]]
        textos = [c.value for fila in hoja.iter_rows() for c in fila if c.value]
        self.assertIn('SE-TECNICO', textos)

    def test_no_dispara_una_consulta_por_activo(self):
        """
        El reporte lista toda la bodega cuando recién se carga (todo en 0 es
        todo crítico), así que una consulta por fila se notaría. Lo evita el
        select_related/prefetch_related de alertas_tecnica.
        """
        from django.db import connection
        from django.test.utils import CaptureQueriesContext

        for i in range(12):
            self.activo(codigo=f'SE-N{i}', existencia=0)

        with CaptureQueriesContext(connection) as consultas:
            for a in reportes.alertas_tecnica():
                _ = (a.nivel_alerta, a.cantidad_afuera,
                     str(a.categoria or ''), str(a.proveedor or ''))

        self.assertLessEqual(
            len(consultas.captured_queries), 4,
            f'{len(consultas.captured_queries)} consultas para 12 activos: falta el prefetch',
        )
